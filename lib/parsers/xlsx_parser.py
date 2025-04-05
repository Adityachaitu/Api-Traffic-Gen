"""Module for parsing xlsx file and create test case json."""
import copy
import os
import random
import re
import time
from urllib.parse import urlencode, quote_plus
import json_flatten
from openpyxl import load_workbook

from lib.converters.xml_convertor import *
from lib.im_environment import ImEnvironment
from lib.label_value_generator import generate_fake_data, mutate_label_value
from lib.parsers import ede_parser
from lib.loggers.logger import report_logger

obj_xml = XMLConvertor()
logger = report_logger()

# logger.logging.getLogger('faker').setLevel(logging.ERROR)


# import lib.loggers.logger


CONFIG_FILE = os.path.join(os.path.dirname(__file__), '../../config/config.yaml')
dynamic_path_param_count = 1
dynamic_path_param_name = ""


def convert_parameter_body(param_json):
    '''
    Reads Request/Response JSON and converts it if it contains EDE metadata.
    Else, it returns JSON as it is.
    :param param_json: JSON as read from excel row
    :return:
    '''
    try:
        param_body = json.loads(param_json)
        if 'ede_info' in param_body:
            ede_param = ede_parser.create_parameter_body(input_json=param_body['ede_info'])
            return ede_param
        else:
            return param_body
    except Exception as e:
        logger.warning("Not a json param. Hence returning as same as the input")
        return param_json

class XlSXParser:
    """Class to parse xlsx file."""

    MANDATORY_FIELDS = {
        "url": False,
        "host": False,
        "method": False,
        "rsp_code": False,
        "req_body": False,
        "rsp_body": False
    }

    INPUT_ELEMENT = {
        "url": "",
        "host": "",
        "method": "",
        "rsp_code": 0,
        "req_body": [],
        "rsp_body": []
    }

    OUTPUT_ELEMENT = {
        "url_group": "",
        "url_present": None,
        "url_instances": [],
        "host": "",
    }

    def __init__(self, file_path, sheet_name, test_env='', test_tag=''):
        """Initialize class object.

        :param file_path: str, path to the xlsx path
        :param sheet_name: str, name of the sheet in the tab
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.test_env = test_env
        self.test_tag = test_tag
        self.ImObj = ImEnvironment(CONFIG_FILE)
        label_data_file = self.ImObj.label_json_file
        if os.path.isfile(label_data_file):
            with open(label_data_file) as f:
                self.label_data = json.load(f)
        else:
            self.label_data = {}

    def return_host(self, host_name):
        '''
        This method converts host in case, host column has config key(e.g. site_name), value should have config. as prefix
        Otherwise, it adds timestamp as suffix and returns it
        :param host_name:
        :return:
        '''
        is_matched = re.search(r"^config\.", host_name)
        if is_matched:
            site_var = host_name[is_matched.end():].lower()
            if self.test_env and 'site_name' in site_var:
                try:
                    host_name = self.ImObj.get_env_param(param=self.test_env)[site_var]
                except:
                    host_name = self.ImObj.get_env_param(param=site_var)
            else:
                host_name = self.ImObj.get_env_param(param=site_var)
            return host_name
        else:
            random_host_suffix = str(time.time_ns())
            return host_name + "." + random_host_suffix

    def _parse_url_dynamic_param_and_enter_value(self, template_url):
        """Method for parsing URL templates and substituting dynamic param path.

        :param template_url: string, request URL strings which has path parameter segments.
        :return URL, with dynamic param path.
        """
        global dynamic_path_param_name
        global dynamic_path_param_count
        delimiter = "/"
        segments = template_url.split(delimiter)
        temp_segment_list = []
        for i in segments:
            if str(i).startswith("<") and str(i).endswith(">"):
                label_name = i.replace("<", "").replace(">", "")
                if label_name.lower() == "dynamic-param":
                    # print(dynamic_path_param_name)
                    label_value = dynamic_path_param_name + str(dynamic_path_param_count)
                    dynamic_path_param_count += 1
                else:
                    label_value = i
            else:
                label_value = i
            temp_segment_list.append(label_value)
        final_url = delimiter.join(temp_segment_list)
        return final_url

    def _test_case_parser(self, file):
        """Method to parse test case from xlsx file.

        :param file: obj, openpyxl workbook object.
        :return test_cases: dict/json,
        output in particular format.
        """
        tag_column = None
        test_case_id = 0
        test_cases = {}
        row_list = []
        input_list = []
        output_list = []
        metadata = {}
        index_to_column_map = {}
        unique_list = []
        req_unique_values = {}
        rsp_unique_values = {}
        global dynamic_path_param_name

        for row in file.iter_rows():
            row_list.append(row)

        column_names = row_list[0]
        row_list_no_headers = row_list[1:]
        for column_index, column in enumerate(column_names):
            column_name = column.value
            if column_index == 0:
                continue
            column_metadata = column_name.split("_")[0].upper()
            actual_column_name = "_".join(
                column_name.split("_")[1:]
            )
            if actual_column_name == "Robot_Tag":  # Storing column index for test tag
                tag_column = column_index
            # print(tag_column)
            if column_metadata == "INPUT":
                index_to_column_map[column_index] = (
                    actual_column_name, column_metadata
                )
            if column_metadata == "OUTPUT":
                index_to_column_map[column_index] = (
                    actual_column_name, column_metadata
                )
            if column_metadata == "META":
                index_to_column_map[column_index] = (
                    actual_column_name, column_metadata
                )
            # Added to read column containing feature specific usecases, should start with feature_
            if column_metadata == "FEATURE":
                index_to_column_map[column_index] = (
                    actual_column_name, column_metadata
                )
            # Added to read column containing list of labels, whose sensitivity should be enabled
            if column_metadata == "SENSITIVE":
                index_to_column_map[column_index] = (
                    actual_column_name, column_metadata
                )
        random_host_suffix = str(time.time_ns())
        mandatory_fields = copy.deepcopy(self.MANDATORY_FIELDS)
        req_body_list = []
        resp_body_list = []
        query_params_list = []
        skip_test = False
        for row_index, row in enumerate(row_list_no_headers):
            if tag_column != None:
                if row[tag_column].value != None and self.test_tag not in row[tag_column].value:
                    skip_test = True
                elif row[tag_column].value != None and self.test_tag in row[tag_column].value:
                    skip_test = False
            if skip_test:
                continue
            url_with_query_params_list = []
            input_element_copy = {}
            output_element_copy = {}
            metadata_element_copy = {}
            # Flags to check if Request body, Response Body are given in the format of SOAP+XML / XML / Json
            flag_req_xml = False
            flag_rsp_xml = False
            flag_soap = False
            mandatory_fields_copy = copy.deepcopy(mandatory_fields)
            mandatory_fields = copy.deepcopy(self.MANDATORY_FIELDS)
            if not any([column.value for column in row]):
                random_host_suffix = str(time.time_ns())
                if len(input_list) > 0 and len(output_list) > 0:
                    output_not_none_list = []
                    for i, ol in enumerate(output_list):
                        if any(ol.values()):
                            output_not_none_list.append(i)
                    output_list_temp = []
                    for i in output_not_none_list:
                        output_list_temp.append(output_list[i])
                    output_list = output_list_temp
                    metadata["payload_structure"] = {
                        "request": None, "response": None, "query_params": None
                    }
                    metadata["payload_structure"]["request"] = req_body_list
                    metadata["payload_structure"]["response"] = resp_body_list
                    metadata["payload_structure"]["query_params"] = query_params_list
                    test_cases[test_case_id] = {
                        "input": input_list,
                        "output": output_list,
                        "metadata": metadata,
                    }
                    for field, present in mandatory_fields_copy.items():
                        if not present:
                            logger.error("Mandatory field " + field + "not present.")
                    input_list = []
                    output_list = []
                    metadata = {}
                    req_body_index = None
                    rsp_body_index = None
                    req_body_list = []
                    resp_body_list = []
                    query_params_list = []
                    unique_list = []
                    req_unique_values = {}
                    rsp_unique_values = {}
                    test_case_id = 0
                    continue
            for column_index, column in enumerate(row):
                if column_index == 0:
                    if column.value:
                        test_case_id = int(column.value)
                        logger.debug("Test Case ID: " + str(test_case_id))
                    continue
                column_name = index_to_column_map[column_index][0]
                column_section = index_to_column_map[column_index][1]
                column_value = column.value

                if column_name.upper() == "URL":
                    if "|" in column_value:
                        url_column_list = column_value.strip().split("|")
                        if len(url_column_list) == 3:
                            event_count = \
                                url_column_list[2].strip().split(":")[1]
                        elif len(url_column_list) == 2:
                            event_count = 1
                        query_params_json = json.loads(
                            url_column_list[1]
                        )
                        query_params_list.append(query_params_json)
                        url_without_query_params = url_column_list[0]
                        for _ in range(int(event_count)):
                            if query_params_json != json.loads("{}"):
                                query_params_json_copy = copy.deepcopy(
                                    query_params_json
                                )
                                '''
                                    Labelling is now done in label_parser
                                '''
                                # labelled_query_params_json = \
                                #     self._label_query_params_json(query_params_json_copy, unique_list)
                                query_params = urlencode(
                                    query_params_json_copy, quote_via=quote_plus
                                )
                                query_params = query_params.replace("%3C","<")
                                query_params = query_params.replace("%3E", ">")
                                url_with_query_params_list.append(
                                    url_without_query_params + "?" + query_params
                                )
                            elif query_params_json == json.loads("{}"):
                                url_with_query_params_list.append(
                                    url_without_query_params
                                )
                    else:
                        url_with_query_params_list = [column_value]
                        event_count = 1
                    mandatory_fields[column_name.lower()] = True

                if column_name.upper() == "HOST":
                    input_element_copy["host"] = \
                        self.return_host(column_value)
                    mandatory_fields[column_name.lower()] = True

                if column_name.upper() == "METHOD":
                    input_element_copy["method"] = column_value
                    mandatory_fields[column_name.lower()] = True

                if column_name.upper() == "RSP_CODE":
                    input_element_copy["rsp_code"] = int(column_value)
                    mandatory_fields[column_name.lower()] = True

                if column_name.upper() == "REQ_HEADER":
                    if column_value:
                        if "application/soap+xml" in column_value.lower():
                            flag_soap = True

                if column_name.upper() == "REQ_BODY":
                    # Now request body would also be able to read EDE metadata, and return payload, along with existing
                    # functionality, this is performed by convert_parameter_body method
                    if column_value:
                        # If the Req.Header is application/soap+xml, check if the Req.Body has the data in XML / Json format
                        if flag_soap:
                            if obj_xml.is_valid_xml(column_value) == True:
                                flag_req_xml = True
                                req_body = convert_parameter_body(column_value)
                                req_body_list.append(req_body)
                            else:
                                req_body = convert_parameter_body(column_value)
                                req_body_list.append(req_body)
                        else:
                            req_body = convert_parameter_body(column_value)
                            req_body_list.append(req_body)
                    req_body_index = column_index

                if column_name.upper() == "RSP_BODY":
                    # Now response body would also be able to read EDE metadata, and return payload, along with existing
                    # functionality, this is performed by convert_parameter_body method
                    if column_value:
                        # If the Req.Header is application/soap+xml, check if the Resp.Body has the data in XML / Json format
                        if flag_soap:
                            if obj_xml.is_valid_xml(column_value) == True:
                                flag_rsp_xml = True
                                resp_body = convert_parameter_body(column_value)
                                resp_body_list.append(resp_body)
                            else:
                                resp_body = convert_parameter_body(column_value)
                                resp_body_list.append(resp_body)
                        else:
                            resp_body = convert_parameter_body(column_value)
                            resp_body_list.append(resp_body)
                    rsp_body_index = column_index

                if column_name.upper() == "SETTINGS" and column_section == "FEATURE":
                    # returns feature settings data from xlsx sheet
                    if column_value:
                        metadata_element_copy["settings"] = json.loads(column_value)

                if column_name.upper() == "VALIDATION" and column_section == "FEATURE":
                    # returns feature validation data from xlsx sheet
                    if column_value:
                        metadata_element_copy["validation"] = json.loads(column_value)

                if column_name.upper() == "DATA_LABELS":
                    # returns sensitive data labels as list from xlsx sheet in labels delimited by ,
                    # (e.g. creditcard,address)
                    if column_value:
                        metadata_element_copy["data_labels"] = column_value.strip().split(",")

                if column_name.upper() == "DYNAMIC_PATH_PARAM_NAME":
                    # returns dynamic param name , (e.g. v,ver)
                    if column_value:
                        dynamic_path_param_name = column_value.strip()

                if column_name.lower() not in mandatory_fields.keys():
                    if column_section == "INPUT":
                        input_element_copy[column_name] = column_value
                    if column_section == "OUTPUT":
                        output_element_copy[column_name] = column_value
                    if column_section == "META":
                        metadata_element_copy[column_name] = column_value
                if column_index == len(row) - 1:
                    for i in range(int(event_count)):
                        input_element_copy["req_body"] = None
                        input_element_copy["rsp_body"] = None
                        if "<" not in url_with_query_params_list[i]:
                            input_element_copy["url"] = url_with_query_params_list[i]
                        else:
                            input_element_copy["url"] = self._parse_url_dynamic_param_and_enter_value(
                                url_with_query_params_list[i])
                        label_data_dict = self.label_data.get(url_with_query_params_list[i], {}).get(
                            input_element_copy["method"], {})
                        if not flag_req_xml:
                            if row[req_body_index].value:
                                '''
                                    Label parsing now done by label_parser
                                '''
                                # labelled_req_body = self._parse_templates_and_enter_value(
                                #     convert_parameter_body(
                                #         row[req_body_index].value
                                #     ), label_data_dict
                                #     , req_unique_values
                                # )
                                input_element_copy["req_body"] = row[req_body_index].value
                                # print(req_unique_values)
                            else:
                                input_element_copy["req_body"] = None
                        else:
                            input_element_copy["req_body"] = row[req_body_index].value

                        if not flag_rsp_xml:
                            if row[rsp_body_index].value:
                                '''
                                    Label parsing now done by label_parser
                                '''
                                # labelled_rsp_body = self._parse_templates_and_enter_value(
                                #     convert_parameter_body(
                                #         row[rsp_body_index].value
                                #     ), label_data_dict
                                #     , rsp_unique_values
                                # )
                                input_element_copy["rsp_body"] = row[rsp_body_index].value
                            else:
                                input_element_copy["rsp_body"] = None
                        else:
                            input_element_copy["rsp_body"] = row[rsp_body_index].value

                        input_element_copy_final = copy.deepcopy(input_element_copy)
                        input_list.append(input_element_copy_final)
                    mandatory_fields["req_body"] = True
                    mandatory_fields["rsp_body"] = True
            output_list.append(output_element_copy)
            if any(metadata_element_copy.values()):
                metadata = copy.deepcopy(metadata_element_copy)

            if row_index == len(row_list_no_headers) - 1:
                if len(input_list) > 0 and len(output_list) > 0:
                    output_not_none_list = []
                    for i, ol in enumerate(output_list):
                        if any(ol.values()):
                            output_not_none_list.append(i)
                    output_list_temp = []
                    for i in output_not_none_list:
                        output_list_temp.append(output_list[i])
                    output_list = output_list_temp
                    metadata["payload_structure"] = {
                        "request": None, "response": None, "query_params": None
                    }
                    metadata["payload_structure"]["request"] = req_body_list
                    metadata["payload_structure"]["response"] = resp_body_list
                    metadata["payload_structure"]["query_params"] = query_params_list
                    test_cases[test_case_id] = {
                        "input": input_list,
                        "output": output_list,
                        "metadata": metadata,
                    }
                    for field, present in mandatory_fields.items():
                        if not present:
                            logger.error("Mandatory field " + field + "not present.")
                    input_list = []
                    output_list = []
                    metadata = {}
                    req_body_index = None
                    rsp_body_index = None
                    req_body_list = []
                    resp_body_list = []
                    query_params_list = []
                    unique_list = []
                    req_unique_values = {}
                    rsp_unique_values = {}
                    test_case_id = 0
        logger.debug("test_cases: %s" %str(test_cases))
        # print(test_cases)
        return test_cases

    def get_formatted_testcases(self):
        """Method to return formatted testcases.

        :return dict/json
        """
        xlsx_file_path = self.file_path
        workbook = load_workbook(xlsx_file_path)
        xlsx_file = workbook[self.sheet_name]
        return self._test_case_parser(xlsx_file)

    def get_json_file(self, json_file):
        test_data = self.get_formatted_testcases()
        file_name = open(json_file, 'w')
        file_name.write(json.dumps(test_data))
        file_name.close()

if __name__ == "__main__":
    try:
        file_to_save, file_to_process, sheet_name, test_env = sys.argv[1:5]
    except:
        file_to_save, file_to_process, sheet_name = sys.argv[1:4]
        test_env = ''
    obj = XlSXParser(file_to_process, sheet_name, test_env)
    obj.get_json_file(file_to_save)
